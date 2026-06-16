# -*- coding: utf-8 -*-
import os
import shutil
import zipfile
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASE_DIR = "/Users/force/Google_Antigravity/horizon_class/skyline-class/class2"

# Mappings of files to copy and rename from original resources
RENAMED_MAPPING = {
    "招標需求/台南玩具展_RFP.pdf": "2026_招標需求_台南玩具展_RFP.pdf",
    "招標需求/台北設計週_RFP.pdf": "2026_招標需求_台北設計週_RFP.pdf",
    "協力商大同/大同搭建商_實績證明.pdf": "2023_大同搭建_大同搭建商_實績證明.pdf",
    "協力商大同/大同搭建商_環保建材認證.pdf": "2023_大同搭建_大同搭建商_環保建材認證.pdf",
    "協力商華麗/華麗設計商_實績證明.pdf": "2024_華麗設計_華麗設計商_實績證明.pdf",
    "協力商名音/名音燈光商_SLA承諾.pdf": "2026_名音燈光_名音燈光商_SLA承諾.pdf",
    "協力商極光/極光音響商_SLA承諾.pdf": "2026_極光音響_極光音響商_SLA承諾.pdf",
    "公司證照庫/陳俊宏_Alex_PMP_2025.pdf": "2025_陳俊宏_Alex_PMP_2025.pdf",
    "公司證照庫/林淑芬_Sophia_PMP_2028.pdf": "2028_林淑芬_Sophia_PMP_2028.pdf",
    "公司證照庫/張雅婷_Tina_PMP_2029.pdf": "2029_張雅婷_Tina_PMP_2029.pdf",
    "公司證照庫/王志明_Jimmy_PMP_2026.pdf": "2026_王志明_Jimmy_PMP_2026.pdf"
}

RFP_TXT_CONTENT = """2026年台南玩具博覽會 ─ 招標需求說明書 (RFP)

一、專案背景與概述
本專案為「2026年台南玩具博覽會」之展位規劃、場地搭建與展期運維技術支援服務。主辦單位旨在透過公開招標，徵選具有豐富展會搭建經驗及卓越售後維運保證的合格廠商。

二、廠商投標基本資歷門檻 (硬性合規限制)
本案設有以下硬性合規審查項目，任一項目未達標準者，將直接判定不合規並予以廢標：
1. 專案經理 (PM) 資歷：投標廠商擬派之專案經理，必須持有國際專案管理學會 (PMI) 頒發之有效專案管理師 (PMP) 證照，且證照有效期限不可早於本案投標日 (2026年6月)。
2. 協力商單案實績金額：本案之大部搭設工程允許委由合格協力商執行。該協力商必須具備類似大型展會搭建經驗，且過去類似案件之單一合約實績金額須達到新台幣 3,000,000 元以上 (含 300 萬)。

三、現場運維服務水平承諾 (SLA 限制)
本博覽會展覽期間（2026年6月1日至6月10日），為確保現場設備運作順暢：
1. 故障響應時間 (SLA)：發生技術故障通知後，協力廠商之現場維運工程師必須於 2 小時內 抵達展會現場完成故障排查與修復。若服務水平協議承諾超過 2 小時者，判定不合規並視同廢標。

四、加分評審項目
1. 環保建材佔比：展位搭建之材料中，若使用環保綠色建材佔比大於 30% 以上者，得於技術審查評分中額外給予加分。
"""

DRAFT_MD_CONTENT = """# 2026年台南玩具博覽會 ─ 服務建議書 (草稿)

## 一、專案執行團隊與成員資歷
本公司高度重視本案，特別指派資深專案經理帶領團隊提供專業展會規劃與維運服務：
- **專案經理**：林淑芬 (Sophia)
- **證照字號**：PMI PMP #1982736 (有效期限至 2028-12-31)
- **展會執行實績**：曾擔任「2023年高雄動漫節大型展位規劃」專案負責人，具備超過五次大型展場卓越管理與協調溝通能力。

## 二、協力廠商與搭建實績
為確保展位結構安全與美觀，本案大部搭設工程特別委由大同空間工程有限公司協同執行：
- **協力商名稱**：大同空間工程有限公司
- **實績合約金額**：新台幣 3,500,000 元整 (符合 RFP 300 萬以上之硬性規定)
- **環保建材宣告**：本案承諾使用綠色環保材料比例達到 35%，高於 RFP 規定之加分門檻。

## 三、展期現場維運服務水平 (SLA)
針對展期運維，本團隊提供高規格快速故障抵達與檢修保證：
- **故障抵達響應時間**：承諾接獲故障通報後，於 2 小時內 (現場派駐或周邊待命) 抵達完成故障排查與備品更換，確保展出不中斷。
[人類審查點: 確證 SLA 溢價補償金 NT$ 300,000 已於財務核算核准。]

## 四、財務報價與毛利核算
- **投標總金額**：新台幣 4,500,000 元整
- **執行成本預估**：新台幣 3,800,000 元整 (內含名音燈光 2h SLA 溢價與合約準備金共 300,000 元)
- **預估利潤率**：15.55% (符合 15.0% 毛利防禦紅線)
[人類審查點: 確證已將此投標金額與決策結果寫入地端真理庫。]
"""

# HTML Report templates helper
def get_report_html(title, body_content):
    return f"""<!DOCTYPE html>
<html lang="zh-TW">
<head>
    <meta charset="UTF-8">
    <title>{title}</title>
    <style>
        body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif; background: #0f172a; color: #e2e8f0; padding: 2rem; margin: 0; }}
        .container {{ max-width: 900px; margin: 0 auto; background: #1e293b; padding: 2rem; border-radius: 12px; border: 1px solid #334155; box-shadow: 0 10px 15px -3px rgba(0, 0, 0, 0.3); }}
        h1 {{ color: #38bdf8; font-size: 1.5rem; margin-top: 0; border-bottom: 2px solid #334155; padding-bottom: 0.75rem; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 1.5rem; }}
        th {{ background: #0f172a; text-align: left; padding: 0.75rem; font-size: 0.9rem; color: #38bdf8; border: 1px solid #334155; }}
        td {{ padding: 0.75rem; font-size: 0.85rem; border: 1px solid #334155; }}
        tr:nth-child(even) {{ background: #1e293b; }}
        tr:nth-child(odd) {{ background: #172554; }}
        .badge {{ display: inline-block; padding: 0.2rem 0.5rem; border-radius: 4px; font-weight: bold; font-size: 0.75rem; }}
        .badge-success {{ background: #065f46; color: #34d399; }}
        .badge-fail {{ background: #7f1d1d; color: #f87171; }}
        .badge-warning {{ background: #78350f; color: #fbbf24; }}
        .footer {{ margin-top: 2rem; font-size: 0.75rem; color: #64748b; text-align: center; border-top: 1px solid #334155; padding-top: 1rem; }}
    </style>
</head>
<body>
    <div class="container">
        {body_content}
        <div class="footer">Falo AIDE 智慧審計控制塔 - 稽核報告存檔</div>
    </div>
</body>
</html>"""

def build_excel_db(db_path, step_num):
    wb = openpyxl.Workbook()
    # Sheet 1: 檔案清冊
    ws_files = wb.active
    ws_files.title = "檔案清冊"
    files_headers = ["file_id", "original_name", "staged_name", "staged_path", "file_size_bytes", "category"]
    ws_files.append(files_headers)
    
    # Sheet 2: 欄位萃取 metadata
    ws_meta = wb.create_sheet(title="欄位萃取 metadata")
    meta_headers = ["meta_id", "file_id", "year_extracted", "company_name", "amount_extracted", "cert_expiry_date", "sla_response_hours"]
    ws_meta.append(meta_headers)

    # Sheet 3: 已得標標案 (For Step 5 historical compare & Step 6 final SSOT)
    if step_num >= 5:
        ws_bids = wb.create_sheet(title="已得標標案")
        bids_headers = ["bid_id", "project_name", "budget", "assigned_pm", "pm_cert_status", "submitted_date", "approved_by", "status"]
        ws_bids.append(bids_headers)
        # Add historical record
        ws_bids.append(["H001", "2023年高雄動漫節大型展位規劃案", 3500000, "Sophia", "ACTIVE", "2023-08-15", "Admin", "Completed"])
        if step_num >= 6:
            ws_bids.append(["B001", "2026年台南玩具博覽會展位搭設與維運案", 4500000, "Sophia", "ACTIVE (2028-12-31)", "2026-06-16", "Force(ff)", "Submitted"])
        
    if step_num >= 6:
        ws_audit = wb.create_sheet(title="審計日誌")
        audit_headers = ["log_id", "timestamp", "action", "operator", "hash_code"]
        ws_audit.append(audit_headers)
        ws_audit.append(["L001", "2026-06-16 19:42:10", "Commit final bidding details to SSOT", "Force", "ff892b10aef"])

    # Base file records (with mock physical sizes)
    file_records = [
        ("F001", "台南玩具展_RFP.pdf", "2026_招標需求_台南玩具展_RFP.pdf", "RFP/招標規範", 15420),
        ("F002", "台北設計週_RFP.pdf", "2026_招標需求_台北設計週_RFP.pdf", "RFP/招標規範", 12450),
        ("F003", "大同搭建商_實績證明.pdf", "2023_大同搭建_大同搭建商_實績證明.pdf", "協力商實績證明", 8450),
        ("F004", "大同搭建商_環保建材認證.pdf", "2023_大同搭建_大同搭建商_環保建材認證.pdf", "協力商環保認證", 5600),
        ("F005", "華麗設計商_實績證明.pdf", "2024_華麗設計_華麗設計商_實績證明.pdf", "協力商實績證明", 7890),
        ("F006", "名音燈光商_SLA承諾.pdf", "2026_名音燈光_名音燈光商_SLA承諾.pdf", "協力商SLA承諾", 6120),
        ("F007", "極光音響商_SLA承諾.pdf", "2026_極光音響_極光音響商_SLA承諾.pdf", "協力商SLA承諾", 5430),
        ("F008", "陳俊宏_Alex_PMP_2025.pdf", "2025_陳俊宏_Alex_PMP_2025.pdf", "專案經理證照", 8120),
        ("F009", "林淑芬_Sophia_PMP_2028.pdf", "2028_林淑芬_Sophia_PMP_2028.pdf", "專案經理證照", 9150),
        ("F010", "張雅婷_Tina_PMP_2029.pdf", "2029_張雅婷_Tina_PMP_2029.pdf", "專案經理證照", 8740),
        ("F011", "王志明_Jimmy_PMP_2026.pdf", "2026_王志明_Jimmy_PMP_2026.pdf", "專案經理證照", 7630),
    ]
    
    for fid, orig, staged, cat, size in file_records:
        ws_files.append([fid, orig, staged, f"標案工作暫存區/{staged}", size, cat])

    # Metadata records - dynamically update values depending on step
    # Step 2 & 3 start: Jimmy is PM, Ming Yin is 4h SLA, wood rates normal
    # Step 4 start: Sophia is RAG retrieved, Ming Yin cost updated to 1.1M and SLA 2h
    # Step 5 & 6 start: Sophia replaced Jimmy, budget synced.
    
    sla_hours = 4 if step_num < 4 else 2
    wood_amount = 3500000
    if step_num >= 4:
        # Step 3.1: wood rates +5% (3.5M -> 3.675M)
        wood_amount = 3675000

    meta_records = [
        ("M001", "F001", 2026, "台南玩具博覽會主辦單位", None, None, None),
        ("M002", "F002", 2026, "台北國際設計週主辦單位", None, None, None),
        ("M003", "F003", 2023, "大同空間工程有限公司", wood_amount, None, None),
        ("M004", "F004", 2023, "大同空間工程有限公司", None, None, None),
        ("M005", "F005", 2024, "華麗展演設計股份有限公司", 2800000, None, None),
        ("M006", "F006", 2026, "名音舞台燈光音響有限公司", None, None, sla_hours),
        ("M007", "F007", 2026, "極光舞台特效音響有限公司", None, None, 2),
        ("M008", "F008", 2025, "Project Management Institute", None, "2025-08-31", None),
        ("M009", "F009", 2028, "Project Management Institute", None, "2028-12-31", None),
        ("M010", "F010", 2029, "Project Management Institute", None, "2029-05-31", None),
        ("M011", "F011", 2026, "Project Management Institute", None, "2026-02-15", None),
    ]
    
    for mid, fid, year, comp, amt, expiry, sla in meta_records:
        ws_meta.append([mid, fid, year, comp, amt, expiry, sla])

    # Design Excel layout
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    sheets_to_format = [ws_files, ws_meta]
    if step_num >= 5:
        sheets_to_format.append(ws_bids)
    if step_num >= 6:
        sheets_to_format.append(ws_audit)

    for ws in sheets_to_format:
        # Format headers
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        ws.row_dimensions[1].height = 24
        
        # Style cells
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).border = thin_border
        
        # Auto-fit column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
            
    wb.save(db_path)

def generate_workspace(step_num):
    part_name = f"part{step_num}"
    part_dir = os.path.join(BASE_DIR, part_name)
    
    # Recreate clean workspace folder (unpacked)
    if os.path.exists(part_dir):
        shutil.rmtree(part_dir)
    os.makedirs(part_dir, exist_ok=True)
    
    # Copy original resources and templates
    shutil.copytree(os.path.join(BASE_DIR, "原始標案資源"), os.path.join(part_dir, "原始標案資源"))
    shutil.copytree(os.path.join(BASE_DIR, "參考範例"), os.path.join(part_dir, "參考範例"))
    shutil.copytree(os.path.join(BASE_DIR, "notebooklm_shared_brains"), os.path.join(part_dir, "notebooklm_shared_brains"))
    
    staging_dir = os.path.join(part_dir, "標案工作暫存區")
    os.makedirs(staging_dir, exist_ok=True)
    
    # Generate files in Staging area depending on step number
    if step_num >= 2:
        # Step 1 finished outputs: renamed files, RFP txt, initial database
        for rel_src, dst_name in RENAMED_MAPPING.items():
            src_path = os.path.join(part_dir, "原始標案資源", rel_src)
            dst_path = os.path.join(staging_dir, dst_name)
            shutil.copy(src_path, dst_path)
            
        with open(os.path.join(staging_dir, "2026_招標需求_台南玩具展_RFP.txt"), "w", encoding="utf-8") as f:
            f.write(RFP_TXT_CONTENT)
            
        build_excel_db(os.path.join(staging_dir, "標案工作記錄表.xlsx"), step_num)
        
    if step_num >= 3:
        # Step 2 finished outputs: HTML reports
        # clause_checklist.html
        body = """<h1>2026年台南玩具博覽會 ─ RFP 廢標條款與合規對帳表</h1>
        <table>
            <tr><th>稽核項目</th><th>RFP 硬性規定說明</th><th>權責單位</th><th>關鍵廢標紅線</th></tr>
            <tr><td>專案經理持證資歷</td><td>擬派專案經理必須持有有效 PMP 證照，有效期限不可早於 2026 年 6 月。</td><td>人力資源部</td><td><span class="badge badge-fail">YES</span></td></tr>
            <tr><td>協力商單案實績</td><td>協力商過去類似案件之單一合約實績金額須達 3,000,000 元以上。</td><td>採購合規組</td><td><span class="badge badge-fail">YES</span></td></tr>
            <tr><td>故障響應時間 (SLA)</td><td>展期故障維修響應，協力商工程師必須於 2 小時內抵達現場排障。</td><td>外協運維部</td><td><span class="badge badge-fail">YES</span></td></tr>
        </table>"""
        with open(os.path.join(staging_dir, "clause_checklist.html"), "w", encoding="utf-8") as f:
            f.write(get_report_html("RFP 廢標條款對帳表", body))
            
        # pm_audit_report.html
        body = """<h1>專案經理 (PM) 證照合規稽核報告</h1>
        <p>目前擬派專案經理：<strong>王志明 (Jimmy)</strong></p>
        <table>
            <tr><th>檢查項目</th><th>RFP 要求</th><th>擬派人員狀態</th><th>合規判定</th></tr>
            <tr><td>證照種類</td><td>PMI PMP 有效證書</td><td>王志明 (PMP #1002934)</td><td><span class="badge badge-success">合規</span></td></tr>
            <tr><td>到期日驗證</td><td>不早於 2026-06</td><td>失效日：2026-02-15 (已過期)</td><td><span class="badge badge-fail">不合規 (RED_DISQUALIFY)</span></td></tr>
            <tr><td>備用替代方案</td><td>尋找證照有效 PM</td><td>證照庫中林淑芬 (Sophia) 證照效期至 2028-12-31，實績符合。</td><td><span class="badge badge-warning">建議啟動備降調撥</span></td></tr>
        </table>"""
        with open(os.path.join(staging_dir, "pm_audit_report.html"), "w", encoding="utf-8") as f:
            f.write(get_report_html("PM 證照合規稽核報告", body))
            
        # sla_gap_report.html
        body = """<h1>協力廠商運維 SLA 差距分析報告</h1>
        <table>
            <tr><th>項目</th><th>RFP 規格要求</th><th>名音燈光商原始承諾</th><th>差距 (Gap)</th><th>合規判定</th></tr>
            <tr><td>SLA現場響應時間</td><td>2 小時以內現場抵達</td><td>4 小時以內抵達現場</td><td>+2 小時 (延遲風險)</td><td><span class="badge badge-warning">不合規 (YELLOW_WARNING)</span></td></tr>
        </table>
        <p><strong>調整方案：</strong> 名音提議增加 300,000 元排班準備金即可將 SLA 上調至 2h；或改為極光音響 (2h SLA)。建議追加預算以保留長期合作名音。</p>"""
        with open(os.path.join(staging_dir, "sla_gap_report.html"), "w", encoding="utf-8") as f:
            f.write(get_report_html("運維 SLA 差距分析報告", body))

    if step_num >= 4:
        # Step 3 finished outputs: Cost adjustments
        # material_adjust_report.html
        body = """<h1>木作搭建材料費率調校報告 (+5%)</h1>
        <table>
            <tr><th>材料項目</th><th>原單價 (NT$)</th><th>調校後單價 (NT$)</th><th>漲幅比例</th><th>影響成本增額</th></tr>
            <tr><td>大同空間搭建工程 (木作與地毯)</td><td>3,500,000</td><td>3,675,000</td><td>5%</td><td>+175,000</td></tr>
        </table>"""
        with open(os.path.join(staging_dir, "material_adjust_report.html"), "w", encoding="utf-8") as f:
            f.write(get_report_html("材料費率調校報告", body))
            
        # sla_cost_report.html
        body = """<h1>名音燈光 2h SLA 溢價與合約準備金評估報告</h1>
        <table>
            <tr><th>費用項目</th><th>SLA 加急規格</th><th>名音溢價報價 (NT$)</th><th>資金分攤比例</th><th>建議核准狀態</th></tr>
            <tr><td>SLA 升級排班加急費</td><td>2h 故障響應</td><td>300,000</td><td>主辦單位與外協 5:5 撥備</td><td><span class="badge badge-success">GREEN_APPROVE</span></td></tr>
        </table>"""
        with open(os.path.join(staging_dir, "sla_cost_report.html"), "w", encoding="utf-8") as f:
            f.write(get_report_html("SLA 溢價評估報告", body))
            
        # profit_analysis_report.html
        body = """<h1>台南玩具展總包標案利潤壓力測試與財務決策書</h1>
        <table>
            <tr><th>項目名稱</th><th>數值 (NT$)</th><th>備註說明</th></tr>
            <tr><td>專案總預算</td><td>4,500,000</td><td>招標書限額</td></tr>
            <tr><td>基礎搭建成本</td><td>3,675,000</td><td>含大同搭建 5% 通膨調幅</td></tr>
            <tr><td>SLA 升級加急溢價</td><td>300,000</td><td>名音燈光升級 2h 現場響應</td></tr>
            <tr><td>專案淨利潤</td><td>525,000</td><td>扣除上述成本後所得</td></tr>
            <tr><td>實質毛利率</td><td><strong>11.67%</strong></td><td>已低於 15.0% 毛利防禦紅線</td></tr>
            <tr><td>決策判定</td><td><span class="badge badge-warning">利潤預警 (YELLOW_WARNING)</span></td><td>若直接承接將面臨高成本壓力，建議優化木作建材降低耗損。</td></tr>
        </table>"""
        with open(os.path.join(staging_dir, "profit_analysis_report.html"), "w", encoding="utf-8") as f:
            f.write(get_report_html("利潤與壓力測試報告", body))

    if step_num >= 5:
        # Step 4 finished outputs: compiled RAG draft
        with open(os.path.join(staging_dir, "draft_assembled.md"), "w", encoding="utf-8") as f:
            f.write(DRAFT_MD_CONTENT)
            
        # retrieval_manifest.html
        body = """<h1>NotebookLM 知識庫真理檢索核對清單</h1>
        <table>
            <tr><th>檢索主題</th><th>擷取來源</th><th>核對核心內容</th><th>資料可信度</th></tr>
            <tr><td>林淑芬 (Sophia) 證照</td><td>公司證照庫/林淑芬_PMP.pdf</td><td>證書號 #1982736, 有效期至 2028-12-31</td><td><span class="badge badge-success">真實有效</span></td></tr>
            <tr><td>歷年大型展會得標實績</td><td>歷史案卷/2023_高雄動漫節.pdf</td><td>大同工程承包, 單案實績金額 3,500,000 元</td><td><span class="badge badge-success">真實有效</span></td></tr>
        </table>"""
        with open(os.path.join(staging_dir, "retrieval_manifest.html"), "w", encoding="utf-8") as f:
            f.write(get_report_html("知識庫真理檢索核對清單", body))
            
        # pm_replacement_log.html
        body = """<h1>專案經理 (PM) 備降替換安全日誌</h1>
        <table>
            <tr><th>原擬任 PM</th><th>更替後新 PM</th><th>證照截止日</th><th>資歷調整說明</th><th>異動狀態</th></tr>
            <tr><td>王志明 (Jimmy)</td><td>林淑芬 (Sophia)</td><td>2028-12-31</td><td>Jimmy 證照已於 2026-02 過期。Sophia 具備大型展會資歷且證照有效。</td><td><span class="badge badge-success">變更完成</span></td></tr>
        </table>"""
        with open(os.path.join(staging_dir, "pm_replacement_log.html"), "w", encoding="utf-8") as f:
            f.write(get_report_html("PM 替換安全日誌", body))
            
        # hitl_review_checklist.html
        body = """<h1>建議書草稿人類在環安全審核檢核表 (HITL)</h1>
        <table>
            <tr><th>段落行號</th><th>審查安全標記</th><th>變更類型</th><th>審查狀態</th></tr>
            <tr><td>Line 25</td><td>[人類審查點: 確證 SLA 溢價補償金 NT$ 300,000 已核准。]</td><td>財務變更</td><td><span class="badge badge-warning">待主管簽章</span></td></tr>
            <tr><td>Line 32</td><td>[人類審查點: 確證已將此投標金額與決策結果寫入地端真理庫。]</td><td>地端寫回</td><td><span class="badge badge-warning">待簽章確認</span></td></tr>
        </table>"""
        with open(os.path.join(staging_dir, "hitl_review_checklist.html"), "w", encoding="utf-8") as f:
            f.write(get_report_html("HITL 安全審核檢核表", body))

    if step_num >= 6:
        # Step 5 finished outputs: Completeness & Gap Analysis
        # draft_completeness_report.html
        body = """<h1>2026年台南玩具博覽會 ─ 建議書完整度評估報告</h1>
        <table>
            <tr><th>對比項目</th><th>玩具展草稿內容</th><th>高雄動漫節歷史規格</th><th>比對結果</th><th>完整度評估</th></tr>
            <tr><td>技術架構與搭設工程</td><td>大同空間工程承包，環保材料佔比 35%</td><td>大同空間工程承包，實績 3.5M</td><td><span class="badge badge-success">技術規格完全承接</span></td><td>100%</td></tr>
            <tr><td>證照合規性</td><td>PM Sophia (PMP 有效期至 2028-12-31)</td><td>PM 證照有效</td><td><span class="badge badge-success">證照效期完全對齊</span></td><td>100%</td></tr>
            <tr><td>財務預算利潤</td><td>報價 4.5M，利潤率 15.55%</td><td>符合毛利防禦紅線</td><td><span class="badge badge-success">高於 15% 紅線</span></td><td>100%</td></tr>
        </table>"""
        with open(os.path.join(staging_dir, "draft_completeness_report.html"), "w", encoding="utf-8") as f:
            f.write(get_report_html("建議書草稿完整度報告", body))
            
        # compliance_gap_analysis.html
        body = """<h1>2026年台南玩具博覽會 ─ 投標前合規差異分析報告</h1>
        <table>
            <tr><th>RFP 合規項目</th><th>點檢狀態</th><th>草稿內對應人類審查點說明</th><th>安全覆核狀態</th></tr>
            <tr><td>專案經理持證資歷</td><td><span class="badge badge-success">符合 (GREEN)</span></td><td>[人類審查點] 已更替過期 PM 志明為有效 PM Sophia</td><td><span class="badge badge-success">主管已確認</span></td></tr>
            <tr><td>SLA 現場響應時間</td><td><span class="badge badge-success">符合 (GREEN)</span></td><td>[人類審查點] 已核准 SLA 30 萬溢價，名音升級 2h SLA</td><td><span class="badge badge-success">主管已確認</span></td></tr>
            <tr><td>協力商單案實績</td><td><span class="badge badge-success">符合 (GREEN)</span></td><td>大同搭建單案 3.5M 符合 300 萬門檻</td><td><span class="badge badge-success">自動通過</span></td></tr>
        </table>"""
        with open(os.path.join(staging_dir, "compliance_gap_analysis.html"), "w", encoding="utf-8") as f:
            f.write(get_report_html("合規差異分析報告", body))

        # Step 6 background logs (for PM Monitoring)
        # pm_schedule_report.html
        body = """<h1>AI PM 排程監控 ─ 關鍵路徑進度風險日誌</h1>
        <table>
            <tr><th>檢查時間</th><th>基準進度</th><th>實際落後時數</th><th>風險等級</th><th>建議調配方案</th></tr>
            <tr><td>2026-06-16 10:00:00</td><td>應完成 RFP 分類與去敏感</td><td>12 小時</td><td><span class="badge badge-warning">DELAY_RISK</span></td></tr>
        </table>"""
        with open(os.path.join(staging_dir, "pm_schedule_report.html"), "w", encoding="utf-8") as f:
            f.write(get_report_html("AI PM 進度風險日誌", body))

        # missing_doc_alert.html
        body = """<h1>外協廠商缺失文件語意催辦報告</h1>
        <table>
            <tr><th>缺失文件</th><th>協力廠商</th><th>預估廢標風險</th><th>催辦郵件窗口</th></tr>
            <tr><td>大同搭建商實績證明.pdf</td><td>大同空間工程有限公司</td><td><span class="badge badge-fail">HIGH_RISK</span></td><td>林大同 (datong@datongspace.com)</td></tr>
        </table>
        <p><strong>催辦信草稿預覽：</strong><br>主旨：【重要合規照會】台南玩具博覽會搭建實績證明補件通知<br>內文：林總您好... 貴司承辦之高雄動漫節實績證明書（350萬元）尚未上傳掃描檔，此文件為招標合規審查之硬性紅線，如未於 6/18 前補件將直接導致本團隊廢標。請速提供以利彙整...</p>"""
        with open(os.path.join(staging_dir, "missing_doc_alert.html"), "w", encoding="utf-8") as f:
            f.write(get_report_html("缺失文件催辦報告", body))

        # cert_recovery_report.html
        body = """<h1>證照過期觸發 PM 備降調配修復報告</h1>
        <table>
            <tr><th>失效人員</th><th>更替人員</th><th>證照比對狀態</th><th>資料庫寫入狀態</th></tr>
            <tr><td>王志明 (過期)</td><td>林淑芬 (有效 2028-12-31)</td><td>PMP 資格符合</td><td><span class="badge badge-success">Success (已同步建議書)</span></td></tr>
        </table>"""
        with open(os.path.join(staging_dir, "cert_recovery_report.html"), "w", encoding="utf-8") as f:
            f.write(get_report_html("PM 備降調配修復報告", body))

    print(f"SUCCESS: Generated unpacked workspace for {part_name} at {part_dir}")

def build_unified_workspace():
    # Build a complete workspace covering Step 6 (final Reference)
    unified_dir = os.path.join(BASE_DIR, "temp_unified")
    if os.path.exists(unified_dir):
        shutil.rmtree(unified_dir)
    os.makedirs(unified_dir, exist_ok=True)
    
    shutil.copytree(os.path.join(BASE_DIR, "原始標案資源"), os.path.join(unified_dir, "原始標案資源"))
    shutil.copytree(os.path.join(BASE_DIR, "參考範例"), os.path.join(unified_dir, "參考範例"))
    shutil.copytree(os.path.join(BASE_DIR, "notebooklm_shared_brains"), os.path.join(unified_dir, "notebooklm_shared_brains"))
    
    staging_dir = os.path.join(unified_dir, "標案工作暫存區")
    os.makedirs(staging_dir, exist_ok=True)
    
    # 1. Copy and rename renamed files
    for rel_src, dst_name in RENAMED_MAPPING.items():
        src_path = os.path.join(unified_dir, "原始標案資源", rel_src)
        dst_path = os.path.join(staging_dir, dst_name)
        shutil.copy(src_path, dst_path)
        
    # 2. Write RFP txt
    with open(os.path.join(staging_dir, "2026_招標需求_台南玩具展_RFP.txt"), "w", encoding="utf-8") as f:
        f.write(RFP_TXT_CONTENT)
        
    # 3. Write draft Suggestion
    with open(os.path.join(staging_dir, "draft_assembled.md"), "w", encoding="utf-8") as f:
        f.write(DRAFT_MD_CONTENT)
        
    # 4. Write Excel DB
    build_excel_db(os.path.join(staging_dir, "標案工作記錄表.xlsx"), 6)
    
    # 5. Write HTML reports
    reports = {
        "clause_checklist.html": "RFP 廢標條款對帳表",
        "pm_audit_report.html": "PM 證照合規稽核報告",
        "sla_gap_report.html": "運維 SLA 差距分析報告",
        "material_adjust_report.html": "材料費率調校報告",
        "sla_cost_report.html": "SLA 溢價評估報告",
        "profit_analysis_report.html": "利潤與壓力測試報告",
        "retrieval_manifest.html": "知識庫真理檢索核對清單",
        "pm_replacement_log.html": "PM 替換安全日誌",
        "hitl_review_checklist.html": "HITL 安全審核檢核表",
        "draft_completeness_report.html": "建議書草稿完整度報告",
        "compliance_gap_analysis.html": "合規差異分析報告",
        "pm_schedule_report.html": "AI PM 進度風險日誌",
        "missing_doc_alert.html": "缺失文件催辦報告",
        "cert_recovery_report.html": "PM 備降調配修復報告"
    }
    
    for filename, title in reports.items():
        with open(os.path.join(staging_dir, filename), "w", encoding="utf-8") as f:
            f.write(get_report_html(title, f"<h1>{title}</h1><p>這是地端已自動生成的審計存檔文件 <strong>{filename}</strong>。</p>"))
            
    # Zip the unified workspace
    zip_path = os.path.join(BASE_DIR, "class2_workspace.zip")
    if os.path.exists(zip_path):
        os.remove(zip_path)
        
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, filenames in os.walk(unified_dir):
            for filename in filenames:
                filepath = os.path.join(root, filename)
                rel_path = os.path.relpath(filepath, unified_dir)
                zipf.write(filepath, rel_path)
                
    print(f"SUCCESS: Created unified class2_workspace.zip at {zip_path}")
    shutil.rmtree(unified_dir)

if __name__ == "__main__":
    # Ensure sub-directories exist
    for i in range(1, 7):
        os.makedirs(os.path.join(BASE_DIR, f"part{i}"), exist_ok=True)
        generate_workspace(i)
        
    build_unified_workspace()
