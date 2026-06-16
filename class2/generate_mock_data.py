import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import letter

# 1. Register macOS Chinese Font
font_path = '/System/Library/Fonts/STHeiti Light.ttc'
if not os.path.exists(font_path):
    font_path = '/System/Library/Fonts/Supplemental/Songti.ttc'
if not os.path.exists(font_path):
    font_path = '/System/Library/Fonts/Helvetica.ttc'  # Fallback

pdfmetrics.registerFont(TTFont('STHeiti', font_path))

# 2. PDF Generator Helper
def create_pdf(filepath, title, paragraphs):
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    doc = SimpleDocTemplate(filepath, pagesize=letter, leftMargin=54, rightMargin=54, topMargin=54, bottomMargin=54)
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'ChineseTitle',
        parent=styles['Heading1'],
        fontName='STHeiti',
        fontSize=16,
        leading=20,
        spaceAfter=15,
        alignment=1  # Center
    )
    
    body_style = ParagraphStyle(
        'ChineseBody',
        parent=styles['Normal'],
        fontName='STHeiti',
        fontSize=10.5,
        leading=16,
        spaceAfter=8
    )
    
    story = []
    story.append(Paragraph(title, title_style))
    story.append(Spacer(1, 10))
    for p in paragraphs:
        story.append(Paragraph(p, body_style))
    doc.build(story)
    print(f"Generated PDF: {filepath}")

# 3. Excel Template Generator Helper
def create_excel_template(filepath):
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb = openpyxl.Workbook()
    
    # Setup Sheet 1: 檔案清冊
    ws_files = wb.active
    ws_files.title = "檔案清冊"
    files_headers = ["file_id", "original_name", "staged_name", "staged_path", "file_size_bytes", "category"]
    ws_files.append(files_headers)
    
    # Setup Sheet 2: 欄位萃取 metadata
    ws_meta = wb.create_sheet(title="欄位萃取 metadata")
    meta_headers = ["meta_id", "file_id", "year_extracted", "company_name", "amount_extracted", "cert_expiry_date", "sla_response_hours"]
    ws_meta.append(meta_headers)
    
    # Apply styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    for ws, headers in [(ws_files, files_headers), (ws_meta, meta_headers)]:
        # Format headers
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        ws.row_dimensions[1].height = 24
        
    # Auto-fit column widths
    for ws in [ws_files, ws_meta]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
            
    wb.save(filepath)
    print(f"Generated Excel Template: {filepath}")

# 4. Generate the Mock Data
if __name__ == "__main__":
    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 4.1 Generate PDFs in "原始標案資源"
    # PDF 1: 台南玩具展_RFP.pdf
    create_pdf(
        os.path.join(base_dir, "原始標案資源", "招標需求", "台南玩具展_RFP.pdf"),
        "2026年台南玩具博覽會 ─ 招標需求說明書 (RFP)",
        [
            "<b>一、專案背景與概述</b>",
            "本專案為「2026年台南玩具博覽會」之展位規劃、場地搭建與展期運維技術支援服務。主辦單位旨在透過公開招標，徵選具有豐富展會搭建經驗及卓越售後維運保證的合格廠商。",
            "<b>二、廠商投標基本資歷門檻 (硬性合規限制)</b>",
            "本案設有以下硬性合規審查項目，任一項目未達標準者，將直接判定不合規並予以廢標：",
            "1. <b>專案經理 (PM) 資歷</b>：投標廠商擬派之專案經理，必須持有國際專案管理學會 (PMI) 頒發之有效專案管理師 (PMP) 證照，且證照有效期限不可早於本案投標日 (2026年6月)。",
            "2. <b>協力商單案實績金額</b>：本案之大部搭設工程允許委由合格協力商執行。該協力商必須具備類似大型展會搭建經驗，且過去類似案件之單一合約實績金額須達到新台幣 <b>3,000,000 元以上</b> (含 300 萬)。",
            "<b>三、現場運維服務水平承諾 (SLA 限制)</b>",
            "本博覽會展覽期間（2026年6月1日至6月10日），為確保現場設備運作順暢：",
            "1. <b>故障響應時間 (SLA)</b>：發生技術故障通知後，協力廠商之現場維運工程師必須於 <b>2 小時內</b> 抵達展會現場完成故障排查與修復。若服務水平協議承諾超過 2 小時者，判定不合規並視同廢標。",
            "<b>四、加分評審項目</b>",
            "1. <b>環保建材佔比</b>：展位搭建之材料中，若使用環保綠色建材佔比大於 30% 以上者，得於技術審查評分中額外給予加分。"
        ]
    )

    # PDF 2: 台北設計週_RFP.pdf
    create_pdf(
        os.path.join(base_dir, "原始標案資源", "招標需求", "台北設計週_RFP.pdf"),
        "2026年台北國際設計週 ─ 招標需求說明書 (RFP)",
        [
            "<b>一、專案背景與概述</b>",
            "本專案為「2026年台北國際設計週」展場空間搭建與多媒體設備租賃服務案。",
            "<b>二、廠商資歷與硬性門檻</b>",
            "1. <b>專案經理 (PM)</b>：必須持有有效 PMP 證照。",
            "2. <b>協力商實績</b>：類似展會搭建單案合約實績金額須達到新台幣 <b>5,000,000 元以上</b> (含 500 萬)。",
            "3. <b>故障響應 SLA</b>：展期內設備發生技術故障，必須於 <b>1 小時內</b> 抵達現場排障，否則視同廢標違約。"
        ]
    )

    # PDF 3: 大同搭建商_實績證明.pdf
    create_pdf(
        os.path.join(base_dir, "原始標案資源", "協力商大同", "大同搭建商_實績證明.pdf"),
        "大同空間工程有限公司 ─ 工程搭建實績證明書",
        [
            "<b>一、專案實績申報資訊</b>",
            "本公司大同空間工程有限公司，曾於2023年承接並順利完工以下工程項目：",
            "1. 專案名稱：2023年高雄動漫節展位設計與搭建工程",
            "2. 實績合約總金額：新台幣 <b>3,500,000 元整</b>",
            "3. 驗收完成時間：2023年10月15日",
            "4. 實績工程內容：負責展場主舞台搭設、各參展商基本展位之隔板、燈具及地毯鋪設工程，並順利通過安全稽核驗收。",
            "<b>二、經辦人與聯絡個資資訊 (機密，僅供合規審查使用)</b>",
            "為配合本年度投標之資歷審查，特此提供當時專案聯絡人之真實身分資訊，以便招標單位照會核對。請審查單位務必妥善處理個資去敏感化防護：",
            "1. 專案現場總負責人：林大同",
            "2. 身分證字號：<b>A123456789</b> (個資敏感項，請遮蔽)",
            "3. 聯絡電話：<b>0912-345-678</b> (個資敏感項，請遮蔽)",
            "4. 公司地址：台北市中山區中山北路三段22號"
        ]
    )

    # PDF 4: 大同搭建商_環保建材認證.pdf
    create_pdf(
        os.path.join(base_dir, "原始標案資源", "協力商大同", "大同搭建商_環保建材認證.pdf"),
        "大同空間工程有限公司 ─ 綠色環保建材標章與佔比證明",
        [
            "<b>一、環保標章資訊</b>",
            "本公司所採購之板材、塗料與鋼架，均符合環保署一類綠建材標準，並領有國家綠建材證書。",
            "<b>二、專案環保建材佔比宣告</b>",
            "針對展位設計，本公司承諾使用之循環回收材料與低揮發性綠色建材佔比達到該展位總材料重量之 <b>35%</b>，高於業界平均標準，符合招標規範之加分門檻要求。"
        ]
    )

    # PDF 5: 華麗設計商_實績證明.pdf
    create_pdf(
        os.path.join(base_dir, "原始標案資源", "協力商華麗", "華麗設計商_實績證明.pdf"),
        "華麗展演設計股份有限公司 ─ 專案實績證明書",
        [
            "<b>一、實績宣告</b>",
            "本公司曾於2024年承辦「台北文創藝術設計展」之空間搭建。",
            "1. 專案名稱：2024年台北文創藝術設計展空間搭設案",
            "2. 實績金額：新台幣 <b>2,800,000 元整</b> (未達 300 萬)",
            "3. 經辦聯絡人：黃華麗 (身分證字號：<b>B234567890</b>, 電話：<b>0922-111-222</b>)"
        ]
    )

    # PDF 6: 名音燈光商_SLA承諾.pdf
    create_pdf(
        os.path.join(base_dir, "原始標案資源", "協力商名音", "名音燈光商_SLA承諾.pdf"),
        "名音舞台燈光音響有限公司 ─ 運維故障響應服務承諾書",
        [
            "<b>一、服務說明與範疇</b>",
            "本公司名音舞台燈光音響有限公司，針對「2026年台南玩具博覽會」之燈光、音響設備租賃及現場技術支援，提供以下服務保證承諾。",
            "<b>二、故障維修與現場響應時間 (SLA) 承諾</b>",
            "本公司承諾於展覽期間（2026年6月1日至6月10日）提供全天候 24 小時電話報修服務：",
            "1. <b>電話初審</b>：接獲報修後 15 分鐘內進行線上初步故障判定。",
            "2. <b>現場抵達響應 (SLA)</b>：若電話排除無效，本公司技術支援小組承諾於接獲通報後 <b>4 小時內</b> 抵達台南玩具展現場，並進行備品更換與故障排除。",
            "<b>三、商務經辦聯絡窗口</b>",
            "1. 專案窗口經辦人：陳名音",
            "2. 聯絡電話：0933-445-566"
        ]
    )

    # PDF 7: 極光音響商_SLA承諾.pdf
    create_pdf(
        os.path.join(base_dir, "原始標案資源", "協力商極光", "極光音響商_SLA承諾.pdf"),
        "極光舞台特效音響有限公司 ─ 現場運維服務水平承諾書 (SLA)",
        [
            "<b>一、SLA 響應條款</b>",
            "為保障展覽設備不間斷運行，極光音響特別針對本案承諾高規格 SLA：",
            "1. <b>故障抵達響應時間</b>：本公司派駐專業工程師於展場週邊待命，一旦接獲技術故障通報，保證於 <b>2 小時內</b> 抵達攤位完成檢修與備用機更換。",
            "<b>二、商務經辦窗口</b>",
            "1. 經辦聯絡人：李極光",
            "2. 聯絡電話：0944-778-899"
        ]
    )

    # PDF 8: 王志明_Jimmy_PMP_2026.pdf
    create_pdf(
        os.path.join(base_dir, "原始標案資源", "公司證照庫", "王志明_Jimmy_PMP_2026.pdf"),
        "Project Management Institute (PMI) - PMP Certificate (Jimmy)",
        [
            "<b>PROJECT MANAGEMENT PROFESSIONAL (PMP)</b>",
            "This is to certify that",
            "<b>Jimmy Wang (王志明)</b>",
            "has successfully met the certification requirements.",
            "PMP Number: 1002934",
            "Original Grant Date: 2020-02-15",
            "Certification Expiration Date: <b>2026-02-15</b> (已過期)",
            "Status: EXPIRED"
        ]
    )

    # PDF 9: 林淑芬_Sophia_PMP_2028.pdf
    create_pdf(
        os.path.join(base_dir, "原始標案資源", "公司證照庫", "林淑芬_Sophia_PMP_2028.pdf"),
        "Project Management Institute (PMI) - PMP Certificate (Sophia)",
        [
            "<b>PROJECT MANAGEMENT PROFESSIONAL (PMP)</b>",
            "This is to certify that",
            "<b>Sophia Lin (林淑芬)</b>",
            "has successfully met the certification requirements.",
            "PMP Number: 1982736",
            "Original Grant Date: 2022-12-31",
            "Certification Expiration Date: <b>2028-12-31</b> (有效)",
            "Status: ACTIVE"
        ]
    )

    # PDF 10: 陳俊宏_Alex_PMP_2025.pdf
    create_pdf(
        os.path.join(base_dir, "原始標案資源", "公司證照庫", "陳俊宏_Alex_PMP_2025.pdf"),
        "Project Management Institute (PMI) - PMP Certificate (Alex)",
        [
            "<b>PROJECT MANAGEMENT PROFESSIONAL (PMP)</b>",
            "This is to certify that",
            "<b>Alex Chen (陳俊宏)</b>",
            "has successfully met the certification requirements.",
            "PMP Number: 1049283",
            "Original Grant Date: 2019-08-31",
            "Certification Expiration Date: <b>2025-08-31</b> (已過期)",
            "Status: EXPIRED"
        ]
    )

    # PDF 11: 張雅婷_Tina_PMP_2029.pdf
    create_pdf(
        os.path.join(base_dir, "原始標案資源", "公司證照庫", "張雅婷_Tina_PMP_2029.pdf"),
        "Project Management Institute (PMI) - PMP Certificate (Tina)",
        [
            "<b>PROJECT MANAGEMENT PROFESSIONAL (PMP)</b>",
            "This is to certify that",
            "<b>Tina Chang (張雅婷)</b>",
            "has successfully met the certification requirements.",
            "PMP Number: 2048591",
            "Original Grant Date: 2023-05-31",
            "Certification Expiration Date: <b>2029-05-31</b> (有效)",
            "Status: ACTIVE"
        ]
    )
    
    # 4.2 Generate Excel template in "參考範例"
    create_excel_template(os.path.join(base_dir, "參考範例", "標案工作記錄表_範本.xlsx"))
    
    print("\nSUCCESS: All mock PDF files and Excel template generated successfully.")
