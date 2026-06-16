# -*- coding: utf-8 -*-
import os
import shutil
import zipfile
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def build_part2_workspace():
    base_dir = "/Users/force/Google_Antigravity/horizon_class/skyline-class/class2"
    part2_dir = os.path.join(base_dir, "part2")
    temp_workspace = os.path.join(part2_dir, "temp_workspace")
    
    # 1. Recreate clean temp workspace
    if os.path.exists(temp_workspace):
        shutil.rmtree(temp_workspace)
    os.makedirs(temp_workspace, exist_ok=True)
    
    # 2. Copy original resources and templates (like part1 structure)
    shutil.copytree(os.path.join(base_dir, "原始標案資源"), os.path.join(temp_workspace, "原始標案資源"))
    shutil.copytree(os.path.join(base_dir, "參考範例"), os.path.join(temp_workspace, "參考範例"))
    
    # Create Staging Area
    staging_dir = os.path.join(temp_workspace, "標案工作暫存區")
    os.makedirs(staging_dir, exist_ok=True)
    
    # 3. Copy and Rename PDFs to represent Step 1.2 smart renaming
    renamed_mapping = {
        "招標需求/台南玩具展_RFP.pdf": "2026_招標需求_台南玩具展_RFP.pdf",
        "招標需求/台北設計週_RFP.pdf": "2026_招標需求_台北設計週_RFP.pdf",
        "協力商大同/大同搭建商_實績證明.pdf": "2023_大同搭建_大同搭建商_實績證明.pdf",
        "協力商大同/大同搭建商_環保建材認證.pdf": "2023_大同搭建_大同搭建商_環保建材認證.pdf",
        "協力商華麗/華麗設計商_實績證明.pdf": "2024_華麗設計_華麗設計商_實績證明.pdf",
        "協力商名音/名音燈光商_SLA承諾.pdf": "2026_名音燈光_名音燈光商_SLA承諾.pdf",
        "協力商極光/極光音響商_SLA承諾.pdf": "2026_極光音響_極光音響商_SLA承諾.pdf",
        "公司證照庫/陳俊宏_Alex_PMP_2025.pdf": "2025_陳俊宏_Alex_PMP_2025.pdf",
        "公司證照庫/林淑芬_Sophia_PMP_2028.pdf": "2028_林淑芬_Sophia_PMP_2028.pdf",
        "公司證照庫/張雅婷_Tina_PMP_2029.pdf": "2029_張雅婷_Tina_PMP_2029.pdf",
        "公司證照庫/王志明_Jimmy_PMP_2026.pdf": "2026_王志明_Jimmy_PMP_2026.pdf"
    }
    
    for rel_src, dst_name in renamed_mapping.items():
        src_path = os.path.join(temp_workspace, "原始標案資源", rel_src)
        dst_path = os.path.join(staging_dir, dst_name)
        shutil.copy(src_path, dst_path)
        print(f"Staged & Renamed: {dst_name}")
        
    # 4. Generate RFP text file (Step 2.1 input)
    rfp_text_path = os.path.join(staging_dir, "2026_招標需求_台南玩具展_RFP.txt")
    rfp_content = """2026年台南玩具博覽會 ─ 招標需求說明書 (RFP)

一、專案背景與概述
本專案為「2026年台南玩具博覽會」之展位規劃、場地搭建與展期運維技術支援服務。主辦單位旨在透過公開招標，徵選具有豐富展會搭建經驗及卓越售後維運保證的合格廠商。

二、廠商投標基本資歷門檻 (硬性合規限制)
本案設有以下硬性合規審查項目，任一項目未達標準者，將直接判定不合規並予以廢標：
1. 專案經理 (PM) 資歷：投標廠商擬派之專案經理，必須持有國際專案管理學會 (PMI) 頒發之有效專案管理師 (PMP) 證照，且證照有效期限不可早於本案投標日 (2026年6月)。
2. 協力商單案實績金額：本案之大部搭設工程允許委由合格協力商執行。該協力商必須具備類似大型展會搭建經驗，且過去類似案件之單一合約實績金額須達到新台幣 3,000,000 元以上 (含 300 萬)。

三、現場運維服務水平承諾 (SLA 限制)
本博覽會展覽期間（2026年6月1日至6月10日），為確保現場設備運作順暢：
1. 故障響應時間 (SLA)：發生技術故障通知後，協力廠商之現場維運工程師必須於 2 小時內 抵達展會現場完成故障排查與修復。若服務水平協議承諾超過 2 小時者，判定不合規並視同廢標。

四、加分評審項目
1. 環保建材佔比：展位搭建之材料中，若使用環保綠色建材佔比大於 30% 以上者，得於技術審查評分中額外給予加分。
"""
    with open(rfp_text_path, "w", encoding="utf-8") as f:
        f.write(rfp_content)
    print("Generated RFP txt file.")

    # 5. Populate Excel Database (標案工作記錄表.xlsx)
    db_path = os.path.join(staging_dir, "標案工作記錄表.xlsx")
    wb = openpyxl.Workbook()
    
    # Sheet 1: 檔案清冊
    ws_files = wb.active
    ws_files.title = "檔案清冊"
    files_headers = ["file_id", "original_name", "staged_name", "staged_path", "file_size_bytes", "category"]
    ws_files.append(files_headers)
    
    # Sheet 2: 欄位萃取 metadata
    ws_meta = wb.create_sheet(title="欄位萃取 metadata")
    meta_headers = ["meta_id", "file_id", "year_extracted", "company_name", "amount_extracted", "cert_expiry_date", "sla_response_hours"]
    ws_meta.append(meta_headers)
    
    # Define File Data Row mappings (with physical size query)
    file_records = [
        ("F001", "台南玩具展_RFP.pdf", "2026_招標需求_台南玩具展_RFP.pdf", "RFP/招標規範"),
        ("F002", "台北設計週_RFP.pdf", "2026_招標需求_台北設計週_RFP.pdf", "RFP/招標規範"),
        ("F003", "大同搭建商_實績證明.pdf", "2023_大同搭建_大同搭建商_實績證明.pdf", "協力商實績證明"),
        ("F004", "大同搭建商_環保建材認證.pdf", "2023_大同搭建_大同搭建商_環保建材認證.pdf", "協力商環保認證"),
        ("F005", "華麗設計商_實績證明.pdf", "2024_華麗設計_華麗設計商_實績證明.pdf", "協力商實績證明"),
        ("F006", "名音燈光商_SLA承諾.pdf", "2026_名音燈光_名音燈光商_SLA承諾.pdf", "協力商SLA承諾"),
        ("F007", "極光音響商_SLA承諾.pdf", "2026_極光音響_極光音響商_SLA承諾.pdf", "協力商SLA承諾"),
        ("F008", "陳俊宏_Alex_PMP_2025.pdf", "2025_陳俊宏_Alex_PMP_2025.pdf", "專案經理證照"),
        ("F009", "林淑芬_Sophia_PMP_2028.pdf", "2028_林淑芬_Sophia_PMP_2028.pdf", "專案經理證照"),
        ("F010", "張雅婷_Tina_PMP_2029.pdf", "2029_張雅婷_Tina_PMP_2029.pdf", "專案經理證照"),
        ("F011", "王志明_Jimmy_PMP_2026.pdf", "2026_王志明_Jimmy_PMP_2026.pdf", "專案經理證照"),
    ]
    
    for fid, orig, staged, cat in file_records:
        full_staged_path = os.path.join(staging_dir, staged)
        fsize = os.path.getsize(full_staged_path)
        ws_files.append([fid, orig, staged, f"標案工作暫存區/{staged}", fsize, cat])
        
    # Metadata Row mappings
    meta_records = [
        ("M001", "F001", 2026, "台南玩具博覽會主辦單位", None, None, None),
        ("M002", "F002", 2026, "台北國際設計週主辦單位", None, None, None),
        ("M003", "F003", 2023, "大同空間工程有限公司", 3500000, None, None),
        ("M004", "F004", 2023, "大同空間工程有限公司", None, None, None),
        ("M005", "F005", 2024, "華麗展演設計股份有限公司", 2800000, None, None),
        ("M006", "F006", 2026, "名音舞台燈光音響有限公司", None, None, 4),
        ("M007", "F007", 2026, "極光舞台特效音響有限公司", None, None, 2),
        ("M008", "F008", 2025, "Project Management Institute", None, "2025-08-31", None),
        ("M009", "F009", 2028, "Project Management Institute", None, "2028-12-31", None),
        ("M010", "F010", 2029, "Project Management Institute", None, "2029-05-31", None),
        ("M011", "F011", 2026, "Project Management Institute", None, "2026-02-15", None),
    ]
    
    for mid, fid, year, comp, amt, expiry, sla in meta_records:
        ws_meta.append([mid, fid, year, comp, amt, expiry, sla])
        
    # Styling matching generate_mock_data.py
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    for ws, headers in [(ws_files, files_headers), (ws_meta, meta_headers)]:
        # Format headers
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        ws.row_dimensions[1].height = 24
        
        # Style cells
        for row in range(2, ws.max_row + 1):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = thin_border
        
        # Auto-fit column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
            
    wb.save(db_path)
    print("Populated and formatted Excel database.")
    
    # 6. Zip the workspace
    zip_path = os.path.join(part2_dir, "part2.zip")
    if os.path.exists(zip_path):
        os.remove(zip_path)
        
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, filenames in os.walk(temp_workspace):
            for filename in filenames:
                filepath = os.path.join(root, filename)
                rel_path = os.path.relpath(filepath, temp_workspace)
                zipf.write(filepath, rel_path)
                
    print(f"SUCCESS: Created part2.zip at {zip_path}")
    
    # Clean up temp workspace folder to keep workspace clean
    shutil.rmtree(temp_workspace)
    print("Cleaned up temp directories.")

if __name__ == "__main__":
    build_part2_workspace()
